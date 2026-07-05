import csv
import mimetypes

from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.parsers import MultiPartParser, JSONParser
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response as DRFResponse
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from forms.models import Form, FormSettings
from .models import Response as Response, UploadedFile, Session
from .email_utils import send_response_notification


def _file_payload(uploaded_file, request, field_lookup=None):
    file_url = ''
    size = 0
    content_type = mimetypes.guess_type(uploaded_file.filename)[0] or 'application/octet-stream'

    if uploaded_file.file:
        file_url = request.build_absolute_uri(uploaded_file.file.url)
        try:
            size = uploaded_file.file.size
        except (OSError, ValueError):
            size = 0

    field_lookup = field_lookup or {}
    return {
        'id': uploaded_file.id,
        'field_id': uploaded_file.field_id,
        'field_label': field_lookup.get(uploaded_file.field_id, uploaded_file.field_id),
        'filename': uploaded_file.filename,
        'url': file_url,
        'size': size,
        'content_type': content_type,
        'uploaded_at': uploaded_file.uploaded_at.isoformat(),
    }


def _response_payload(response_obj, request, field_lookup=None):
    files = [_file_payload(file_obj, request, field_lookup) for file_obj in response_obj.files.all()]
    return {
        'id': response_obj.id,
        'data_json': response_obj.data_json,
        'answers': response_obj.data_json,
        'submitted_at': response_obj.submitted_at.isoformat(),
        'email': response_obj.email,
        'ip_address': response_obj.ip_address or '',
        'time_to_complete': response_obj.time_to_complete,
        'files': files,
        'uploaded_files': files,
    }


def _field_value(response_obj, field):
    data = response_obj.data_json or {}
    value = data.get(field.get('id'), data.get(field.get('label'), ''))
    if isinstance(value, list):
        return ', '.join(str(item) for item in value)
    if isinstance(value, dict):
        return str(value)
    return value


# ── Public: Upload file (called before submit) ────────────────────────────────
@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser])
def upload_file(request):
    slug     = request.data.get('form_slug', '')
    field_id = request.data.get('field_id', '')
    file_obj = request.FILES.get('file')

    if not file_obj or not slug:
        return DRFResponse({'error': 'Missing file or form_slug.'}, status=400)

    try:
        form = Form.objects.get(slug=slug)
    except Form.DoesNotExist:
        return DRFResponse({'error': 'Form not found.'}, status=404)

    uploaded = UploadedFile.objects.create(
        form=form, field_id=field_id,
        file=file_obj, filename=file_obj.name,
    )

    payload = _file_payload(uploaded, request)
    payload['file_id'] = uploaded.id
    return DRFResponse(payload, status=201)


# ── Public: Submit response ───────────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([AllowAny])
def submit_response(request):
    try:
        
        slug             = request.data.get('form_slug', '')
        data             = request.data.get('data', {})
        fingerprint      = request.data.get('fingerprint', '')
        time_to_complete = request.data.get('time_to_complete', 0)
        file_ids         = request.data.get('file_ids', {})   # {field_id: uploaded_file_id}

        if not slug:
            return DRFResponse({'error': 'form_slug is required.'}, status=400)

        try:
            form = Form.objects.select_related('settings').get(slug=slug)
        except Form.DoesNotExist:
            return DRFResponse({'error': 'Form not found.'}, status=404)

        ok, reason = form.is_accepting()
        if not ok:
            return DRFResponse({'error': reason}, status=403)

        # s = form.settings
        from forms.models import FormSettings
        s, _ = FormSettings.objects.get_or_create(form=form)

        # Duplicate checks
        if s.submission_limit_mode == 'browser' and not fingerprint:
            return DRFResponse({'error': 'This form allows one response per browser. Please enable browser storage and try again.'}, status=400)

        if s.submission_limit_mode == 'browser' and fingerprint:
            if Response.objects.filter(form=form, fingerprint=fingerprint).exists():
                return DRFResponse({'error': 'You have already submitted this form.'}, status=409)

        # Robustly detect an email value in the submitted data. Support keys like
        # 'email', 'Email Address', 'Email', etc., and also accept values that
        # look like an email address even if the key is custom.
        email = ''
        if isinstance(data, dict):
            for k, v in data.items():
                if not v:
                    continue
                if isinstance(v, (list, dict)):
                    # skip complex types
                    continue
                ks = str(k).lower()
                vs = str(v).strip()
                # key contains 'email' is highest priority
                if 'email' in ks:
                    email = vs
                    break
                # fallback: value looks like an email address
                if '@' in vs and '.' in vs.split('@')[-1]:
                    email = vs
                    # don't break yet; prefer a key that contains 'email' if present later
        email = (email or '').strip()
        # normalize
        if email:
            email = email.lower()

        if s.submission_limit_mode == 'email' and not email:
            return DRFResponse({'error': 'This form allows one response per email. Please include an email address.'}, status=400)

        if s.submission_limit_mode == 'email' and email:
            if Response.objects.filter(form=form, email__iexact=email).exists():
                return DRFResponse({'error': 'A response from this email already exists.'}, status=409)

        response_obj = Response.objects.create(
            form=form, data_json=data, fingerprint=fingerprint,
            email=email, time_to_complete=time_to_complete,
            ip_address=request.META.get('REMOTE_ADDR'),
        )

        # Link any uploaded files
        if file_ids:
            for field_id, file_id in file_ids.items():
                UploadedFile.objects.filter(id=file_id, form=form).update(response=response_obj)

        # Send email notification (non-blocking)
        # send_response_notification(form, response_obj)

        return DRFResponse({
            'message': s.confirmation_message,
            'id': response_obj.id,
            'redirect_url': s.redirect_url or '',
        }, status=201)
        
    except Exception as e:
        import traceback
        print("SUBMIT ERROR:", e)
        print(traceback.format_exc())

        return DRFResponse(
            {
                "error": str(e),
                "trace": traceback.format_exc()
            },
            status=500
        )


# ── Public: Session tracking ──────────────────────────────────────────────────
@api_view(['POST'])
@permission_classes([AllowAny])
def track_session(request):
    slug      = request.data.get('form_slug', '')
    session_id = request.data.get('session_id')
    status_val = request.data.get('status', 'started')
    last_step  = request.data.get('last_step', 0)
    last_field = request.data.get('last_field_id', '')

    try:
        form = Form.objects.get(slug=slug)
    except Form.DoesNotExist:
        return DRFResponse({'error': 'Form not found.'}, status=404)

    if session_id:
        session = Session.objects.filter(id=session_id, form=form).first()
        if session:
            session.status = status_val
            session.last_step = last_step
            session.last_field_id = last_field
            session.save()
            return DRFResponse({'session_id': session.id})

    session = Session.objects.create(
        form=form, status=status_val, last_step=last_step, last_field_id=last_field,
    )
    return DRFResponse({'session_id': session.id}, status=201)


# ── Admin: List responses ─────────────────────────────────────────────────────
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def list_responses(request, slug):
    try:
        form = Form.objects.get(slug=slug, owner=request.user)
    except Form.DoesNotExist:
        return DRFResponse({'error': 'Form not found.'}, status=404)

    responses = Response.objects.filter(form=form).prefetch_related('files')
    field_lookup = {
        field.get('id'): field.get('label', field.get('id'))
        for field in (form.fields_json or [])
        if field.get('id')
    }
    data = [_response_payload(r, request, field_lookup) for r in responses]
    return DRFResponse(data)


# ── Admin: Edit response ──────────────────────────────────────────────────────
@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
def edit_response(request, slug, pk):
    try:
        form = Form.objects.get(slug=slug, owner=request.user)
        r    = Response.objects.get(id=pk, form=form)
    except (Form.DoesNotExist, Response.DoesNotExist):
        return DRFResponse({'error': 'Not found.'}, status=404)

    new_data = request.data.get('data_json', {})
    if new_data:
        r.data_json = new_data
        r.save()
    return DRFResponse(_response_payload(r, request))


# ── Admin: Delete one response ────────────────────────────────────────────────
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def delete_response(request, slug, pk):
    try:
        form = Form.objects.get(slug=slug, owner=request.user)
        Response.objects.get(id=pk, form=form).delete()
        return DRFResponse(status=204)
    except (Form.DoesNotExist, Response.DoesNotExist):
        return DRFResponse({'error': 'Not found.'}, status=404)


# ── Admin: Clear all responses ────────────────────────────────────────────────
@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def clear_responses(request, slug):
    try:
        form = Form.objects.get(slug=slug, owner=request.user)
        Response.objects.filter(form=form).delete()
        return DRFResponse({'message': 'All responses cleared.'})
    except Form.DoesNotExist:
        return DRFResponse({'error': 'Form not found.'}, status=404)


# ── Admin: Export Excel ───────────────────────────────────────────────────────
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_excel(request, slug):
    try:
        form = Form.objects.get(slug=slug, owner=request.user)
    except Form.DoesNotExist:
        return DRFResponse({'error': 'Form not found.'}, status=404)

    responses = Response.objects.filter(form=form).prefetch_related('files').order_by('submitted_at')
    fields    = [f for f in (form.fields_json or []) if f.get('type') != 'section']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Responses'

    hfill = PatternFill('solid', fgColor='7C3AED')
    hfont = Font(bold=True, color='FFFFFF', name='Calibri', size=11)

    headers = ['#', 'Submitted At', 'Email', 'IP Address', 'Completion Time (seconds)'] + [f['label'] for f in fields] + ['Uploaded Files']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill; cell.font = hfont
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[cell.column_letter].width = max(16, len(h) + 4)
    ws.row_dimensions[1].height = 22

    for row_num, r in enumerate(responses, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=r.submitted_at.strftime('%d-%m-%Y %H:%M'))
        ws.cell(row=row_num, column=3, value=r.email)
        ws.cell(row=row_num, column=4, value=r.ip_address or '')
        ws.cell(row=row_num, column=5, value=r.time_to_complete)
        for col, field in enumerate(fields, 6):
            val = _field_value(r, field)
            ws.cell(row=row_num, column=col, value=str(val) if val else '')
        file_links = [
            request.build_absolute_uri(file_obj.file.url)
            for file_obj in r.files.all()
            if file_obj.file
        ]
        ws.cell(row=row_num, column=len(headers), value=', '.join(file_links))

    safe_title = ''.join(c for c in form.title if c.isalnum() or c in ' _-')[:30]
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{safe_title}_responses.xlsx"'
    wb.save(resp)
    return resp


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_csv(request, slug):
    try:
        form = Form.objects.get(slug=slug, owner=request.user)
    except Form.DoesNotExist:
        return DRFResponse({'error': 'Form not found.'}, status=404)

    responses = Response.objects.filter(form=form).prefetch_related('files').order_by('submitted_at')
    fields = [f for f in (form.fields_json or []) if f.get('type') != 'section']
    headers = ['#', 'Submitted At', 'Email', 'IP Address', 'Completion Time (seconds)'] + [f['label'] for f in fields] + ['Uploaded Files']

    safe_title = ''.join(c for c in form.title if c.isalnum() or c in ' _-')[:30]
    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = f'attachment; filename="{safe_title}_responses.csv"'

    writer = csv.writer(resp)
    writer.writerow(headers)
    for index, response_obj in enumerate(responses, 1):
        file_links = [
            request.build_absolute_uri(file_obj.file.url)
            for file_obj in response_obj.files.all()
            if file_obj.file
        ]
        row = [
            index,
            response_obj.submitted_at.isoformat(),
            response_obj.email,
            response_obj.ip_address or '',
            response_obj.time_to_complete,
        ]
        row.extend(_field_value(response_obj, field) for field in fields)
        row.append(', '.join(file_links))
        writer.writerow(row)

    return resp
